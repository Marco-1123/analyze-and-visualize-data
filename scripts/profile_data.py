#!/usr/bin/env python3
"""Profile a fixed CSV, JSON, or XLSX dataset into a compact JSON report."""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import json
import math
import re
import statistics
import sys
from collections import Counter
from pathlib import Path
from typing import Any, Iterable


NULL_STRINGS = {"", "null", "none", "n/a", "na", "-", "--"}
DATE_PATTERNS = (
    "%Y-%m-%d",
    "%Y/%m/%d",
    "%Y-%m-%d %H:%M:%S",
    "%Y/%m/%d %H:%M:%S",
    "%Y-%m",
    "%Y/%m",
)


def normalize_cell(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        if stripped.lower() in NULL_STRINGS:
            return None
        return stripped
    return value


def dedupe_headers(headers: Iterable[Any]) -> list[str]:
    seen: Counter[str] = Counter()
    result: list[str] = []
    for index, header in enumerate(headers):
        base = str(header).strip() if header not in (None, "") else f"column_{index + 1}"
        seen[base] += 1
        result.append(base if seen[base] == 1 else f"{base}_{seen[base]}")
    return result


def load_csv(path: Path) -> tuple[list[dict[str, Any]], str]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        sample = handle.read(4096)
        handle.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample)
        except csv.Error:
            dialect = csv.excel
        reader = csv.reader(handle, dialect)
        raw_headers = next(reader, [])
        headers = dedupe_headers(raw_headers)
        rows = [
            {header: normalize_cell(value) for header, value in zip(headers, raw_row)}
            for raw_row in reader
        ]
    return rows, "csv"


def load_json(path: Path) -> tuple[list[dict[str, Any]], str]:
    data = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(data, dict):
        for key in ("rows", "data", "records", "items"):
            if isinstance(data.get(key), list):
                data = data[key]
                break
    if not isinstance(data, list):
        raise ValueError("JSON must contain an array of records or a rows/data/records/items array")
    rows: list[dict[str, Any]] = []
    for index, item in enumerate(data):
        if not isinstance(item, dict):
            raise ValueError(f"JSON record {index} is not an object")
        rows.append({str(key): normalize_cell(value) for key, value in item.items()})
    return rows, "json"


def load_xlsx(path: Path, sheet_name: str | None) -> tuple[list[dict[str, Any]], str]:
    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise RuntimeError(
            "XLSX profiling requires openpyxl. Use the bundled Codex Python runtime or the spreadsheets skill."
        ) from error
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
    iterator = sheet.iter_rows(values_only=True)
    headers = dedupe_headers(next(iterator, []))
    rows = [
        {header: normalize_cell(value) for header, value in zip(headers, raw_row)}
        for raw_row in iterator
        if any(value not in (None, "") for value in raw_row)
    ]
    return rows, f"xlsx:{sheet.title}"


def parse_number(value: Any) -> float | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        number = float(value)
        return number if math.isfinite(number) else None
    if not isinstance(value, str):
        return None
    cleaned = value.strip().replace(",", "")
    multiplier = 1.0
    if cleaned.endswith("%"):
        cleaned = cleaned[:-1]
        multiplier = 0.01
    cleaned = re.sub(r"^[¥￥$€£]\s*", "", cleaned)
    try:
        number = float(cleaned) * multiplier
        return number if math.isfinite(number) else None
    except ValueError:
        return None


def parse_date(value: Any) -> str | None:
    if isinstance(value, dt.datetime):
        return value.isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    if not isinstance(value, str):
        return None
    for pattern in DATE_PATTERNS:
        try:
            parsed = dt.datetime.strptime(value.strip(), pattern)
            return parsed.isoformat()
        except ValueError:
            continue
    return None


def infer_type(non_null: list[Any]) -> str:
    if not non_null:
        return "unknown"
    number_count = sum(parse_number(value) is not None for value in non_null)
    date_count = sum(parse_date(value) is not None for value in non_null)
    bool_count = sum(
        isinstance(value, bool)
        or (isinstance(value, str) and value.lower() in {"true", "false", "yes", "no", "是", "否"})
        for value in non_null
    )
    threshold = max(1, math.ceil(len(non_null) * 0.8))
    if bool_count >= threshold:
        return "boolean"
    if date_count >= threshold:
        return "date"
    if number_count >= threshold:
        return "number"
    return "text"


def percentile(sorted_values: list[float], p: float) -> float | None:
    if not sorted_values:
        return None
    index = (len(sorted_values) - 1) * p
    lower = math.floor(index)
    upper = math.ceil(index)
    if lower == upper:
        return sorted_values[lower]
    return sorted_values[lower] + (sorted_values[upper] - sorted_values[lower]) * (index - lower)


def profile_column(name: str, values: list[Any], row_count: int) -> dict[str, Any]:
    non_null = [normalize_cell(value) for value in values if normalize_cell(value) is not None]
    inferred = infer_type(non_null)
    unique_rendered = {json.dumps(value, ensure_ascii=False, default=str) for value in non_null}
    result: dict[str, Any] = {
        "name": name,
        "type": inferred,
        "nonNull": len(non_null),
        "missing": row_count - len(non_null),
        "missingRate": round((row_count - len(non_null)) / row_count, 6) if row_count else 0,
        "unique": len(unique_rendered),
    }
    if inferred == "number":
        numbers = sorted(number for value in non_null if (number := parse_number(value)) is not None)
        if numbers:
            result["stats"] = {
                "min": numbers[0],
                "p25": percentile(numbers, 0.25),
                "median": statistics.median(numbers),
                "mean": statistics.fmean(numbers),
                "p75": percentile(numbers, 0.75),
                "max": numbers[-1],
                "sum": math.fsum(numbers),
            }
    elif inferred == "date":
        dates = sorted(parsed for value in non_null if (parsed := parse_date(value)) is not None)
        if dates:
            result["range"] = {"min": dates[0], "max": dates[-1]}
    else:
        top = Counter(str(value) for value in non_null).most_common(8)
        result["topValues"] = [{"value": value, "count": count} for value, count in top]
    return result


def profile(rows: list[dict[str, Any]], source: str, input_path: Path) -> dict[str, Any]:
    headers: list[str] = []
    for row in rows:
        for key in row:
            if key not in headers:
                headers.append(key)
    row_count = len(rows)
    columns = [
        profile_column(header, [row.get(header) for row in rows], row_count)
        for header in headers
    ]
    duplicate_rows = row_count - len(
        {
            json.dumps(row, ensure_ascii=False, sort_keys=True, default=str)
            for row in rows
        }
    )
    return {
        "source": {
            "file": input_path.name,
            "format": source,
        },
        "shape": {"rows": row_count, "columns": len(headers)},
        "duplicateRows": duplicate_rows,
        "columns": columns,
        "warnings": [
            warning
            for warning in (
                "dataset is empty" if row_count == 0 else None,
                "duplicate rows detected" if duplicate_rows else None,
                "one or more columns exceed 40% missing values"
                if any(column["missingRate"] > 0.4 for column in columns)
                else None,
            )
            if warning
        ],
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Profile a CSV, JSON, or XLSX dataset.")
    parser.add_argument("--input", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    parser.add_argument("--sheet", help="XLSX sheet name; defaults to the first sheet")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    path = args.input.resolve()
    try:
        suffix = path.suffix.lower()
        if suffix in {".csv", ".tsv"}:
            rows, source = load_csv(path)
        elif suffix == ".json":
            rows, source = load_json(path)
        elif suffix in {".xlsx", ".xlsm"}:
            rows, source = load_xlsx(path, args.sheet)
        else:
            raise ValueError(f"unsupported input format: {suffix}")
        result = profile(rows, source, path)
        output = args.output.resolve()
        output.parent.mkdir(parents=True, exist_ok=True)
        output.write_text(json.dumps(result, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    except (OSError, ValueError, RuntimeError, json.JSONDecodeError) as error:
        print(f"profile failed: {error}", file=sys.stderr)
        return 1
    print(f"profiled: {path}")
    print(f"output: {args.output.resolve()}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
